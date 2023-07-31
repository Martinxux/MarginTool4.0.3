#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2023/7/18 20:20
# @Author  : Martin
# @FileName: write_to_excel_pre.py
# @Software: PyCharm

import openpyxl
from openpyxl.styles import Font
import re


def write_to_excel_pre(data, excel_path):
    # 打开 Excel 文件
    wb = openpyxl.Workbook()
    ws = wb.active

    # 写入数据
    for i, row in enumerate(data):
        for j, value in enumerate(row):
            ws.cell(row=i + 1, column=j + 1, value=value)
    # 设置红色字体
    red_font = Font(color="FF0000", size=11, bold=True, name='微软雅黑')

    # 判断关键字并设置红色字体
    # 从第三行第一列到第五列开始循环
    for row in ws.iter_rows(min_row=3, min_col=1, max_col=5):
        lane_num = row[0].value  # lane_num是第三行第一列的数据
        width = row[1].value
        height = row[2].value
        area = row[3].value
        fom = row[4].value

        # 判断是Gen4还是Gen5,不同的Spec，Fail的数据标红
        if "S" in lane_num and re.match(r"S[01]D0T[01234567]P[01234567]L[01234567]", lane_num):
            if width and float(width) <= 0.51:
                row[1].font = red_font
            if height and float(height) <= 50:
                row[2].font = red_font

        elif "D" in lane_num and re.match(r"D[01234567]T[01234567]P[01234567]L[01234567]", lane_num):
            if width and float(width) < 0.59:
                row[1].font = red_font
            if height and float(height) < 58:
                row[2].font = red_font

        # 由于第一列中有其他字符，需要跳过
        else:
            continue
    # 保存 Excel 文件
    wb.save(excel_path)
