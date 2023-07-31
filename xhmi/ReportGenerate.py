#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2023/5/10 16:14
# @Author  : Martin
# @FileName: compare.py
# @Software: PyCharm

import datetime
import fnmatch
import openpyxl
import os
from openpyxl.styles import Side, Border


def Report(folder_path):
    # 修改目录work_dir
    work_dir = folder_path
    file_name = '*.xlsx'

    # 创建一个新的工作表
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.create_sheet('Data', 0)

    # 设置字体、大小和加粗属性
    bold_font = openpyxl.styles.Font(name='微软雅黑', size=11, bold=True)
    bold_font1 = openpyxl.styles.Font(name='微软雅黑', size=11, bold=False, color="FFFFFF")

    # 遍历所有excel文件的sheet,存为list
    wb_list = []
    sheet_list = []
    # 使用 os.walk() 递归遍历文件夹及其子文件夹中的 Excel 文件
    for root, dirs, files in os.walk(folder_path):
        for f in files:
            if fnmatch.fnmatch(f, file_name):
                wb = openpyxl.load_workbook(os.path.join(root, f))
                wb_list.append(wb)

    # 复制所有sheet中的数据，先只复制值，不复制格式
    for i, wb in enumerate(wb_list):
        sheet = wb.active
        for row in sheet.iter_rows():
            new_row = []
            for source_cell in row:
                new_row.append(source_cell.value)
            new_sheet.append(new_row)

        # 在每个 Excel 文件结束时插入指定数量的空行
        if i < len(wb_list) - 1:
            for k in range(1):  # 插入1行空行，可根据需要调整
                new_sheet.append([])

    # 设置单元格内容居中
    center_alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    # 定义边框样式
    border_style = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))
    # 设置单元格样式
    for row in new_sheet.rows:
        for cell in row:
            cell.font = bold_font
            cell.alignment = center_alignment

    # 合并单元格
    for row_num in [1, 23, 45, 67, 89, 111, 133, 155]:
        new_sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        # 设置单元格样式
        # 设置单元格背景填充为红色
        red_fill = openpyxl.styles.PatternFill(patternType='solid', start_color='800000')
        for cell in new_sheet[row_num]:
            if cell.value:
                cell.font = bold_font1
                cell.fill = red_fill
                cell.border = border_style  # 添加边框

    for row_num in [21, 43, 65, 87, 109, 131, 153, 175]:
        new_sheet.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=4)
        # 设置单元格背景填充为绿色
        green_fill = openpyxl.styles.PatternFill(patternType='solid', start_color='00FF00')
        for cell in new_sheet[row_num][1:4]:
            cell.border = border_style  # 添加边框
            if cell.value:
                cell.font = bold_font
                cell.fill = green_fill

    for row_num in [2, 24, 46, 68, 90, 112, 134, 156]:
        # 设置单元格背景填充为粉色
        pink_fill = openpyxl.styles.PatternFill(patternType='solid', start_color='FF8080')
        for cell in new_sheet[row_num]:
            if cell.value:
                cell.font = bold_font
                cell.fill = pink_fill

    # 使用列表解析语句定义需要设置样式的行号列表
    row_nums = [i for j in [[*range(3, 22)], [*range(25, 44)], [*range(47, 66)], [*range(69, 88)],
                            [*range(91, 110)], [*range(113, 132)], [*range(135, 154)], [*range(157, 176)]] for i in j]

    # 针对行号列表中每个行号，设置其第一列单元格的字体加粗和粉色填充
    pink_fill = openpyxl.styles.PatternFill(patternType='solid', start_color='FF8080')
    for row_num in row_nums:
        cell = new_sheet.cell(row=row_num, column=1)
        if cell.value:
            cell.font = bold_font
            cell.fill = pink_fill

    for col in new_sheet.columns:
        for cell in col:
            if cell.value:
                # calculate width based on length of value
                width = max([len(str(cell.value)) * 1.2, 22])
                new_sheet.column_dimensions[cell.column_letter].width = width
    # 定义边框样式
    border_style = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))

    # 为非空单元格添加边框
    for row in new_sheet.rows:
        for cell in row:
            if cell.value:  # 判断单元格是否为空
                cell.border = border_style  # 添加边框样式

    # 遍历所有单元格，将数字格式化为保留两位小数
    for row in new_sheet.rows:
        for cell in row:
            if isinstance(cell.value, (int, float)):  # 判断单元格是否为数字类型
                cell.number_format = '0.00'

    # 遍历所有单元格，判断B列和C列的值是否需要修改字体颜色
    for row in new_sheet.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            # 如果单元格属于合并单元格，则跳过该单元格
            if any(cell.coordinate in rng for rng in cell.parent.merged_cells.ranges):
                continue

            # 获取数据类型为数值类型或字符串类型的值
            value = cell.value if cell.data_type in ('n', 's') else None

            # 判断 B 列的值是否小于 0.3，如果是则设置字体颜色为红色
            if cell.column_letter == 'B' and isinstance(value, float) and value < 0.3:
                cell.font = openpyxl.styles.Font(name='微软雅黑', size=11, bold=True, color="FF0000")

            # 判断 C 列的值是否小于 35，如果是则设置字体颜色为红色
            if cell.column_letter == 'C' and isinstance(value, float) and value < 35:
                cell.font = openpyxl.styles.Font(name='微软雅黑', size=11, bold=True, color="FF0000")

    # 保存新的Excel文件
    today = datetime.date.today().strftime("%Y%m%d")
    # 使用当前日期来命名 Excel 文件
    new_wb.save(os.path.join(work_dir, f'XHMI_Margin_SI Test Report_{today}.xlsx'))
