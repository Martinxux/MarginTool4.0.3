#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2023/5/18 19:26
# @Author  : Martin
# @FileName: write_to_excel.py
# @Software: PyCharm
import openpyxl
import os


def write_to_excel_xhmi(data, excel_file_path):
    wb = openpyxl.Workbook()
    ws = wb.active

    # 写入表头
    excel_folder = os.path.dirname(os.path.abspath(excel_file_path))
    ws['A1'] = os.path.basename(excel_folder)
    ws['A2'] = 'Lane'
    ws['B2'] = 'Width(UI)'
    ws['C2'] = 'Height(mv)'
    ws['D2'] = 'Area(Units)'
    ws['A19'] = 'Worst Case (all lanes)'
    ws['A20'] = 'Spec.'
    ws['B20'] = '0.30'
    ws['C20'] = '35.00'
    ws['D20'] = 'NA'
    ws['A21'] = 'Severity'
    ws['B21'] = 'Low'

    # 写入数据
    for i, row in enumerate(data):
        for j, value in enumerate(row):
            if isinstance(value, str) and value.replace(".", "", 1).isdigit():  # 判断是否为数字字符串
                ws.cell(row=i + 3, column=j + 1, value=float(value))
            else:
                ws.cell(row=i + 3, column=j + 1, value=value)

    # 保存Excel文件
    wb.save(excel_file_path)
