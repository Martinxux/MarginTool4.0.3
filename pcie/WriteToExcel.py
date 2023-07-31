#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2023/5/11 11:19
# @Author  : xuhui
# @FileName: WriteToExcel.py
# @Software: PyCharm
import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import re


def write_to_excell(data, excel_file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    # 写入表头
    ws['A1'] = 'Test'
    ws['B1'] = os.path.basename(os.path.dirname(excel_file_path))
    ws['A2'] = 'LaneNum'
    ws['B2'] = 'Width(UI)'
    ws['C2'] = 'Height(mv)'
    ws['D2'] = 'Area(Units)'

    # 写入数据前的数据转换
    for i, row in enumerate(data):
        converted_row = []
        for value in row:
            # 将数据转换为数字类型
            if isinstance(value, str) and value.replace('.', '').isdigit():
                value = float(value)
            converted_row.append(value)
        data[i] = converted_row
    # 写入数据
    for i, row in enumerate(data):
        for j, value in enumerate(row):
            ws.cell(row=i + 3, column=j + 1, value=value)
    # 设置单元格样式
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    fill_gray = PatternFill(patternType="solid", start_color="FF8080")
    font_bold = Font(size=11, bold=True, name='微软雅黑')  # 字体大小，加粗，字体名称，字体颜色
    border_thick = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                          bottom=Side(style='thin'))

    # 获取当前Excel的最后一行
    last_row = ws.max_row

    # 添加新行
    ws.cell(row=last_row + 1, column=1, value="Worst Case (all lanes)")

    # 计算最小值
    min_width = min(ws.cell(row=row, column=2).value for row in range(3, last_row + 1))
    min_height = min(ws.cell(row=row, column=3).value for row in range(3, last_row + 1))
    min_area = min(ws.cell(row=row, column=4).value for row in range(3, last_row + 1))
    # 将最小值写入新行
    ws.cell(row=last_row + 1, column=2, value=min_width)
    ws.cell(row=last_row + 1, column=3, value=min_height)
    ws.cell(row=last_row + 1, column=4, value=min_area)

    # 在遍历前设置标志
    added_s_spec = False
    added_d_spec = False
    for row in ws.iter_rows(min_row=3, min_col=1, max_col=4):
        # 判断怎么添加SPec
        lane_num = row[0].value
        if "S" in lane_num and re.match(r"S[01]D0T[01234567]P[01234567]L[01234567]", lane_num):
            if not added_s_spec:
                # 在最后一行下面加一行，内容为：Spec、0.51、50
                ws.append(['HygonSpec.', '＞0.51', '＞50', 'NA'])
                last_row += 1
                added_s_spec = True
                for cell in ws[last_row + 1]:
                    cell.fill = fill_gray
        if "D" in lane_num and re.match(r"D[01234567]T[01234567]P[01234567]L[01234567]", lane_num):
            if not added_d_spec:
                # 在最后一行下面加一行，内容为：Spec、0.59、58
                ws.append(['HygonSpec.', '≥0.59', '≥58', 'NA'])
                last_row += 1
                added_d_spec = True
                for cell in ws[last_row + 1]:
                    cell.fill = fill_gray
        else:
            continue

    ws.append(['Severity', 'Low'])
    # 合并单元格
    ws.merge_cells(start_row=last_row + 2, start_column=2, end_row=last_row + 2, end_column=4)

    merged_cell = ws.cell(row=last_row + 2, column=2)
    fill = PatternFill(start_color="008800", end_color="008800", fill_type="solid")
    merged_cell.fill = fill

    # 设置表头样式和颜色
    for j in range(1, 5):
        cell = ws.cell(row=2, column=j)
        cell.alignment = align_center
        cell.font = font_bold
        cell.fill = fill_gray
        cell.border = border_thick
    # 自适应列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            cell.font = font_bold
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # 设置红色字体
    red_font = Font(color="FF0000", size=11, bold=True, name='微软雅黑')
    # 判断关键字并设置红色字体
    for row in ws.iter_rows(min_row=3, min_col=1, max_col=4):
        lane_num = row[0].value
        width = row[1].value
        height = row[2].value
        area = row[3].value
        # 设置边框和居中
        for cell in row[1:4]:
            cell.border = border_thick
            cell.alignment = align_center

        # 判断是Gen4还是Gen5,不同的Spec，Fail的数据标红
        if "S" in lane_num and re.match(r"S[01]D0T[01234567]P[01234567]L[01234567]", lane_num):
            if width and float(width) <= 0.51:
                row[1].font = red_font
            if height and float(height) <= 50:
                row[2].font = red_font

        if "D" in lane_num and re.match(r"D[01234567]T[01234567]P[01234567]L[01234567]", lane_num):
            if width and float(width) < 0.59:
                row[1].font = red_font
            if height and float(height) < 58:
                row[2].font = red_font

        # 填充颜色
        lane_num_fill = PatternFill(start_color="FF8080", end_color="FF8080", fill_type="solid")
        row[0].fill = lane_num_fill
        row[0].border = border_thick  # 为单元格添加边框
        row[0].alignment = align_center

    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B1'].font = Font(name='微软雅黑', size=11, bold=False, color="FFFFFF")
    ws['B1'].fill = PatternFill(patternType="solid", start_color="880000")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = Font(name='微软雅黑', size=11, bold=False, color="FFFFFF")
    ws['A1'].fill = PatternFill(patternType="solid", start_color="880000")
    ws['A1'].border = border_thick
    ws['B1'].border = border_thick
    # 保存Excel文件
    wb.save(excel_file_path)
